"""
ABECIS 裂縫智慧分析報告產生器
流程：ABECIS 預測圖 → Claude Vision 結構化分析 → Excel 報告（含圖表）

輸出 Excel 工作表：
  1. 裂縫詳細分析   - 逐張影像的 LLM 分析結果
  2. 統計摘要       - 嚴重程度、類型、修補優先級統計
  3. 修補優先清單   - 依緊急程度排序

使用方式：
  cd H:\ChihleeMaster\dev\ABECIS_MODEL_SWAP
  G:\conda\envs\CrackSeg\python.exe abecis_report_generator.py [--test N]
"""
import sys, os, base64, io, json, argparse, time
sys.stdout.reconfigure(encoding='utf-8')

# 載入 .env
for _p in [os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env'),
           r'H:\ChihleeMaster\dev\ABECIS_MODEL_SWAP\.env']:
    if os.path.exists(_p):
        with open(_p) as _f:
            for _l in _f:
                _l = _l.strip()
                if _l and not _l.startswith('#') and '=' in _l:
                    k, v = _l.split('=', 1)
                    if k.strip() not in os.environ:
                        os.environ[k.strip()] = v.strip()
        break

import numpy as np
from pathlib import Path
from PIL import Image, ImageDraw, ImageFont
import anthropic
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers)
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter

# ── 路徑設定 ─────────────────────────────────────────────────────────────────
PRED_DIR = r'H:\ChihleeMaster\abecis_predictions'
RGB_DIR  = r'H:\ChihleeMaster\dev\ABECIS_MODEL_SWAP\concreteCrackSegmentationDataset\rgb'
SPLITS   = r'H:\ChihleeMaster\dev\ABECIS_MODEL_SWAP\data\splits\test.txt'
OUTPUT   = r'H:\ChihleeMaster\dev\ABECIS_MODEL_SWAP\outputs\crack_inspection_report.xlsx'

parser = argparse.ArgumentParser()
parser.add_argument('--test', type=int, default=0, help='只處理前 N 張（0=全部）')
args = parser.parse_args()

api_key = os.environ.get('ANTHROPIC_API_KEY', '')
if not api_key:
    print('ERROR: 找不到 ANTHROPIC_API_KEY，請確認 .env 檔案')
    sys.exit(1)

client = anthropic.Anthropic(api_key=api_key)

# ── 讀取測試集 ────────────────────────────────────────────────────────────────
with open(SPLITS) as f:
    test_ids = [l.strip() for l in f if l.strip()]
if args.test > 0:
    test_ids = test_ids[:args.test]

rgb_index = {p.stem.lower(): p for p in Path(RGB_DIR).iterdir()
             if p.suffix.lower() in ('.jpg','.jpeg','.png')}

Path(OUTPUT).parent.mkdir(parents=True, exist_ok=True)

# ── 建立覆疊影像（原圖 + 紅色半透明裂縫遮罩）──────────────────────────────────
def make_overlay(img_id: str) -> Image.Image | None:
    pred_path = Path(PRED_DIR) / f'{img_id}_pred.png'
    rgb_p = rgb_index.get(img_id.lower())
    if not pred_path.exists() or rgb_p is None:
        return None
    rgb = Image.open(rgb_p).convert('RGB')
    mask_pil = Image.open(pred_path).convert('L')
    if mask_pil.size != rgb.size:
        mask_pil = mask_pil.resize(rgb.size, Image.NEAREST)
    mask = np.array(mask_pil) >= 128
    overlay = np.array(rgb).copy()
    overlay[mask, 0] = np.clip(overlay[mask, 0].astype(int) * 0.4 + 220, 0, 255).astype(np.uint8)
    overlay[mask, 1] = (overlay[mask, 1] * 0.4).astype(np.uint8)
    overlay[mask, 2] = (overlay[mask, 2] * 0.4).astype(np.uint8)
    result = Image.fromarray(overlay)
    # 縮小至 800px 寬以節省 API tokens
    max_w = 800
    if result.width > max_w:
        ratio = max_w / result.width
        result = result.resize((max_w, int(result.height * ratio)), Image.LANCZOS)
    return result

def img_to_b64(pil_img: Image.Image) -> str:
    buf = io.BytesIO()
    pil_img.save(buf, format='JPEG', quality=80)
    return base64.standard_b64encode(buf.getvalue()).decode()

# ── LLM 分析函數 ──────────────────────────────────────────────────────────────
PROMPT = """你是一位混凝土結構檢測專家，正在審查自動裂縫偵測系統的輸出結果。
圖片中紅色區域為自動系統偵測到的裂縫位置。

請分析此圖片並以 JSON 格式回傳以下資訊（只回傳 JSON，不要其他文字）：

{
  "crack_detected": true/false,
  "severity": "無裂縫" | "輕微" | "中度" | "嚴重",
  "coverage_pct": 0~100（估計裂縫覆蓋面積百分比，整數）,
  "crack_type": "無" | "髮絲裂縫" | "結構性裂縫" | "網狀裂縫" | "混合型",
  "orientation": "無" | "水平" | "垂直" | "斜向" | "網狀/不規則",
  "crack_count": 估計可見裂縫條數（整數）,
  "urgency": "無需處理" | "定期監測" | "近期修補" | "緊急處理",
  "description": "一句話描述裂縫狀況（繁體中文，20字以內）"
}

嚴重程度判斷標準：
- 輕微：僅有細小髮絲裂縫，寬度 < 0.2mm
- 中度：明顯裂縫，寬度 0.2~1mm 或多條裂縫
- 嚴重：大裂縫，寬度 > 1mm，或網狀裂縫，或結構性裂縫"""

def analyze_crack(img_id: str, overlay: Image.Image) -> dict:
    try:
        resp = client.messages.create(
            model='claude-haiku-4-5',
            max_tokens=300,
            messages=[{
                'role': 'user',
                'content': [
                    {'type': 'image',
                     'source': {'type': 'base64',
                                'media_type': 'image/jpeg',
                                'data': img_to_b64(overlay)}},
                    {'type': 'text', 'text': PROMPT}
                ]
            }]
        )
        text = resp.content[0].text.strip()
        # 清除可能的 markdown code block
        if text.startswith('```'):
            text = '\n'.join(text.split('\n')[1:-1])
        return json.loads(text)
    except json.JSONDecodeError as e:
        print(f'  JSON parse error [{img_id}]: {e}')
        return None
    except Exception as e:
        print(f'  API error [{img_id}]: {e}')
        return None

# ── 主處理迴圈 ────────────────────────────────────────────────────────────────
print(f'Processing {len(test_ids)} images with Claude Vision...\n')
results = []

for idx, img_id in enumerate(test_ids):
    overlay = make_overlay(img_id)
    if overlay is None:
        print(f'[{idx+1:3d}/{len(test_ids)}] {img_id}  SKIP (no prediction)')
        continue

    data = analyze_crack(img_id, overlay)
    if data is None:
        data = {'crack_detected': True, 'severity': '未知', 'coverage_pct': 0,
                'crack_type': '未知', 'orientation': '未知', 'crack_count': 0,
                'urgency': '定期監測', 'description': '分析失敗'}

    data['img_id'] = img_id
    results.append(data)

    sev = data.get('severity', '?')
    urg = data.get('urgency', '?')
    pct = data.get('coverage_pct', 0)
    print(f'[{idx+1:3d}/{len(test_ids)}] {img_id}  嚴重:{sev}  覆蓋:{pct}%  緊急:{urg}')

print(f'\n分析完成：{len(results)} 張')

# ── 建立 Excel 報告 ───────────────────────────────────────────────────────────
print('\n建立 Excel 報告...')
wb = openpyxl.Workbook()

# ── 樣式定義 ──────────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill('solid', fgColor='2E5EAA')
HEADER_FONT  = Font(bold=True, color='FFFFFF', size=11)
TITLE_FONT   = Font(bold=True, size=14, color='1F3864')
SUBTTL_FONT  = Font(bold=True, size=12, color='2E5EAA')
CENTER       = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT         = Alignment(horizontal='left', vertical='center', wrap_text=True)
THIN         = Side(style='thin', color='CCCCCC')
BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

SEV_FILLS = {
    '無裂縫':   PatternFill('solid', fgColor='D9EAD3'),
    '輕微':     PatternFill('solid', fgColor='FFF2CC'),
    '中度':     PatternFill('solid', fgColor='FCE5CD'),
    '嚴重':     PatternFill('solid', fgColor='F4CCCC'),
    '未知':     PatternFill('solid', fgColor='EFEFEF'),
}
URG_FILLS = {
    '無需處理': PatternFill('solid', fgColor='D9EAD3'),
    '定期監測': PatternFill('solid', fgColor='FFF2CC'),
    '近期修補': PatternFill('solid', fgColor='FCE5CD'),
    '緊急處理': PatternFill('solid', fgColor='F4CCCC'),
    '未知':     PatternFill('solid', fgColor='EFEFEF'),
}

def set_header(ws, row, col, text, width=None):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = HEADER_FILL; c.font = HEADER_FONT
    c.alignment = CENTER; c.border = BORDER
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width

def set_cell(ws, row, col, value, fill=None, font=None, align=None):
    c = ws.cell(row=row, column=col, value=value)
    c.border = BORDER
    c.alignment = align or CENTER
    if fill: c.fill = fill
    if font: c.font = font
    return c

# ══════════════════════════════════════════════════════════════════
# 工作表 1：裂縫詳細分析
# ══════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = '裂縫詳細分析'
ws1.sheet_view.showGridLines = False
ws1.row_dimensions[1].height = 30
ws1.row_dimensions[2].height = 22

# 標題
ws1.merge_cells('A1:I1')
t = ws1['A1']
t.value = '混凝土裂縫自動檢測分析報告  ｜  ABECIS + LLM 視覺分析'
t.font = TITLE_FONT; t.alignment = CENTER
t.fill = PatternFill('solid', fgColor='D0E4F7')

# 欄標題
headers = ['圖片編號', '裂縫偵測', '嚴重程度', '覆蓋面積(%)',
           '裂縫類型', '裂縫走向', '裂縫條數', '修補優先級', '狀況描述']
widths  = [12, 10, 12, 14, 14, 16, 10, 14, 28]
for ci, (h, w) in enumerate(zip(headers, widths), 1):
    set_header(ws1, 2, ci, h, w)

# 資料列
for ri, r in enumerate(results, 3):
    ws1.row_dimensions[ri].height = 18
    set_cell(ws1, ri, 1, r['img_id'])
    detected = '✓ 是' if r.get('crack_detected') else '✗ 否'
    set_cell(ws1, ri, 2, detected)
    sev = r.get('severity', '未知')
    set_cell(ws1, ri, 3, sev, fill=SEV_FILLS.get(sev))
    set_cell(ws1, ri, 4, r.get('coverage_pct', 0))
    set_cell(ws1, ri, 5, r.get('crack_type', '未知'))
    set_cell(ws1, ri, 6, r.get('orientation', '未知'))
    set_cell(ws1, ri, 7, r.get('crack_count', 0))
    urg = r.get('urgency', '未知')
    set_cell(ws1, ri, 8, urg, fill=URG_FILLS.get(urg))
    set_cell(ws1, ri, 9, r.get('description', ''), align=LEFT)

ws1.freeze_panes = 'A3'

# ══════════════════════════════════════════════════════════════════
# 工作表 2：統計摘要 + 圖表
# ══════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('統計摘要')
ws2.sheet_view.showGridLines = False

ws2.merge_cells('A1:F1')
t2 = ws2['A1']
t2.value = '檢測統計摘要'
t2.font = TITLE_FONT; t2.alignment = CENTER
t2.fill = PatternFill('solid', fgColor='D0E4F7')

# ── 嚴重程度統計 ──────────────────────────────────────────────────
sev_counts = {'無裂縫': 0, '輕微': 0, '中度': 0, '嚴重': 0, '未知': 0}
for r in results:
    sev_counts[r.get('severity', '未知')] = sev_counts.get(r.get('severity', '未知'), 0) + 1
sev_counts = {k: v for k, v in sev_counts.items() if v > 0}

ws2.merge_cells('A3:C3')
ws2['A3'].value = '嚴重程度分布'; ws2['A3'].font = SUBTTL_FONT; ws2['A3'].alignment = CENTER
set_header(ws2, 4, 1, '嚴重程度', 14)
set_header(ws2, 4, 2, '張數', 10)
set_header(ws2, 4, 3, '佔比(%)', 12)
total = len(results)
sev_order = ['無裂縫', '輕微', '中度', '嚴重', '未知']
sev_row_start = 5
for ri, k in enumerate(sev_order, sev_row_start):
    if k in sev_counts:
        cnt = sev_counts[k]
        set_cell(ws2, ri, 1, k, fill=SEV_FILLS.get(k))
        set_cell(ws2, ri, 2, cnt)
        set_cell(ws2, ri, 3, round(cnt/total*100, 1))

# ── 嚴重程度長條圖 ──────────────────────────────────────────────────
chart1 = BarChart()
chart1.type = 'col'
chart1.title = '裂縫嚴重程度分布'
chart1.y_axis.title = '張數'
chart1.x_axis.title = '嚴重程度'
chart1.style = 10
chart1.width = 15; chart1.height = 10
sev_data_rows = [i for i, k in enumerate(sev_order, sev_row_start) if k in sev_counts]
if sev_data_rows:
    min_r, max_r = min(sev_data_rows), max(sev_data_rows)
    data_ref = Reference(ws2, min_col=2, min_row=min_r, max_row=max_r)
    cats_ref = Reference(ws2, min_col=1, min_row=min_r, max_row=max_r)
    chart1.add_data(data_ref)
    chart1.set_categories(cats_ref)
    chart1.series[0].title = None
    colors = {'無裂縫':'70AD47','輕微':'FFD966','中度':'F6B26B','嚴重':'E06666','未知':'CCCCCC'}
    for i, k in enumerate([k for k in sev_order if k in sev_counts]):
        dp = DataPoint(idx=i)
        dp.spPr = None
        chart1.series[0].dPt.append(dp)
ws2.add_chart(chart1, 'E3')

# ── 修補優先級統計 ──────────────────────────────────────────────────
urg_counts = {'無需處理': 0, '定期監測': 0, '近期修補': 0, '緊急處理': 0}
for r in results:
    k = r.get('urgency', '定期監測')
    urg_counts[k] = urg_counts.get(k, 0) + 1
urg_counts = {k: v for k, v in urg_counts.items() if v > 0}

row_offset = sev_row_start + len(sev_counts) + 2
ws2.merge_cells(f'A{row_offset}:C{row_offset}')
ws2[f'A{row_offset}'].value = '修補優先級分布'
ws2[f'A{row_offset}'].font = SUBTTL_FONT; ws2[f'A{row_offset}'].alignment = CENTER
set_header(ws2, row_offset+1, 1, '優先級', 14)
set_header(ws2, row_offset+1, 2, '張數', 10)
set_header(ws2, row_offset+1, 3, '佔比(%)', 12)
urg_order = ['無需處理', '定期監測', '近期修補', '緊急處理']
urg_row_start = row_offset + 2
for ri, k in enumerate(urg_order, urg_row_start):
    if k in urg_counts:
        cnt = urg_counts[k]
        set_cell(ws2, ri, 1, k, fill=URG_FILLS.get(k))
        set_cell(ws2, ri, 2, cnt)
        set_cell(ws2, ri, 3, round(cnt/total*100, 1))

# ── 修補優先級圓餅圖 ──────────────────────────────────────────────────
chart2 = PieChart()
chart2.title = '修補優先級分布'
chart2.style = 10
chart2.width = 15; chart2.height = 10
urg_data_rows = [i for i, k in enumerate(urg_order, urg_row_start) if k in urg_counts]
if urg_data_rows:
    min_r2, max_r2 = min(urg_data_rows), max(urg_data_rows)
    data_ref2 = Reference(ws2, min_col=2, min_row=min_r2, max_row=max_r2)
    cats_ref2 = Reference(ws2, min_col=1, min_row=min_r2, max_row=max_r2)
    chart2.add_data(data_ref2)
    chart2.set_categories(cats_ref2)
ws2.add_chart(chart2, 'E18')

# ── 整體統計摘要框 ──────────────────────────────────────────────────
sum_row = urg_row_start + len(urg_counts) + 2
ws2.merge_cells(f'A{sum_row}:C{sum_row}')
ws2[f'A{sum_row}'].value = '整體統計'
ws2[f'A{sum_row}'].font = SUBTTL_FONT; ws2[f'A{sum_row}'].alignment = CENTER

stats_items = [
    ('總檢測張數', total),
    ('有裂縫張數', sum(1 for r in results if r.get('crack_detected'))),
    ('裂縫偵測率', f"{sum(1 for r in results if r.get('crack_detected'))/total*100:.1f}%"),
    ('平均覆蓋面積', f"{np.mean([r.get('coverage_pct',0) for r in results]):.1f}%"),
    ('需緊急處理', urg_counts.get('緊急處理', 0)),
    ('需近期修補', urg_counts.get('近期修補', 0)),
]
for i, (k, v) in enumerate(stats_items, sum_row+1):
    set_header(ws2, i, 1, k, 14)
    set_cell(ws2, i, 2, v)

ws2.column_dimensions['A'].width = 14
ws2.column_dimensions['B'].width = 10
ws2.column_dimensions['C'].width = 12

# ══════════════════════════════════════════════════════════════════
# 工作表 3：修補優先清單（按緊急程度排序）
# ══════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('修補優先清單')
ws3.sheet_view.showGridLines = False

ws3.merge_cells('A1:G1')
ws3['A1'].value = '修補優先清單（依緊急程度排序）'
ws3['A1'].font = TITLE_FONT; ws3['A1'].alignment = CENTER
ws3['A1'].fill = PatternFill('solid', fgColor='D0E4F7')

headers3 = ['優先順序', '圖片編號', '嚴重程度', '覆蓋面積(%)', '裂縫類型', '修補優先級', '狀況描述']
widths3  = [12, 12, 12, 14, 14, 14, 30]
for ci, (h, w) in enumerate(zip(headers3, widths3), 1):
    set_header(ws3, 2, ci, h, w)

urg_priority = {'緊急處理': 0, '近期修補': 1, '定期監測': 2, '無需處理': 3, '未知': 4}
sorted_results = sorted(results,
    key=lambda r: (urg_priority.get(r.get('urgency','未知'), 4),
                   -r.get('coverage_pct', 0)))

for ri, r in enumerate(sorted_results, 3):
    ws3.row_dimensions[ri].height = 18
    urg = r.get('urgency', '未知')
    sev = r.get('severity', '未知')
    fill = URG_FILLS.get(urg)
    set_cell(ws3, ri, 1, ri-2, fill=fill)
    set_cell(ws3, ri, 2, r['img_id'], fill=fill)
    set_cell(ws3, ri, 3, sev, fill=SEV_FILLS.get(sev))
    set_cell(ws3, ri, 4, r.get('coverage_pct', 0), fill=fill)
    set_cell(ws3, ri, 5, r.get('crack_type', '未知'), fill=fill)
    set_cell(ws3, ri, 6, urg, fill=fill)
    set_cell(ws3, ri, 7, r.get('description', ''), fill=fill, align=LEFT)

ws3.freeze_panes = 'A3'

# ── 儲存 ──────────────────────────────────────────────────────────────────────
wb.save(OUTPUT)
print(f'\n✅ 報告已儲存：{OUTPUT}')

# ── 控制台摘要 ────────────────────────────────────────────────────────────────
print(f'\n{"="*55}')
print(f'📊 檢測摘要（{total} 張影像）')
print(f'{"="*55}')
print(f'{"嚴重程度":<12}  {"張數":>5}  {"佔比":>7}')
print('-'*28)
for k in sev_order:
    if k in sev_counts:
        print(f'{k:<12}  {sev_counts[k]:>5}  {sev_counts[k]/total*100:>6.1f}%')
print(f'\n{"修補優先級":<12}  {"張數":>5}  {"佔比":>7}')
print('-'*28)
for k in urg_order:
    if k in urg_counts:
        print(f'{k:<12}  {urg_counts[k]:>5}  {urg_counts[k]/total*100:>6.1f}%')
print(f'\n🚨 緊急處理：{urg_counts.get("緊急處理",0)} 張')
print(f'🔧 近期修補：{urg_counts.get("近期修補",0)} 張')
print(f'📁 輸出檔案：{OUTPUT}')
